from __future__ import annotations

from collections.abc import Iterable
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .config import ConfigError
from .config_schema import BlankMode, FIELD_SPECS, FieldSpec, field_spec_for_key


DATA_SHEETS = (
    "Instructions",
    "Config Parameters",
    "Sampling Channels",
    "Error Sources",
    "Assumptions",
    "Voltage Dip Tests",
)

CONFIG_HEADERS = (
    "Section",
    "Config Key",
    "Parameter Name",
    "Default Value",
    "User Value",
    "Data Type",
    "Unit",
    "Required",
    "Applies When",
    "Allowed Values",
    "Description",
    "Chinese Notes",
)

SAMPLING_HEADERS = (
    "Name",
    "Quantity",
    "ADC Module",
    "SOC",
    "Trigger Source",
    "Aperture Delay (s)",
    "Chinese Notes",
)

ERROR_HEADERS = (
    "Channel",
    "Name",
    "Enabled",
    "Percent FS",
    "Drift (ppm/degC)",
    "Chinese Notes",
)

ASSUMPTION_HEADERS = (
    "Topic",
    "Value",
    "Source Type",
    "Qualification",
    "Chinese Notes",
)

DIP_HEADERS = (
    "Voltage Percent",
    "Duration Cycles 50Hz",
    "Duration Cycles 60Hz",
    "Functional Status",
    "Chinese Notes",
)

_SHEET_HEADERS = {
    "Config Parameters": CONFIG_HEADERS,
    "Sampling Channels": SAMPLING_HEADERS,
    "Error Sources": ERROR_HEADERS,
    "Assumptions": ASSUMPTION_HEADERS,
    "Voltage Dip Tests": DIP_HEADERS,
}
_MISSING = object()
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


def _set_dotted(data: dict[str, Any], key: str, value: Any) -> None:
    current = data
    tokens = key.split(".")
    for token in tokens[:-1]:
        child = current.get(token)
        if child is None:
            child = {}
            current[token] = child
        if not isinstance(child, dict):
            raise ConfigError(f"Cannot set nested configuration key: {key}")
        current = child
    current[tokens[-1]] = value


def _get_dotted(data: dict[str, Any], key: str, default: Any = _MISSING) -> Any:
    current: Any = data
    for token in key.split("."):
        if not isinstance(current, dict) or token not in current:
            return default
        current = current[token]
    return current


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _reject_formula(value: Any, location: str) -> None:
    if isinstance(value, str) and value.startswith("="):
        raise ConfigError(f"[{location}] formula cells are not supported.")


def _parse_scalar(spec: FieldSpec, value: Any, location: str) -> Any:
    _reject_formula(value, location)
    try:
        if spec.data_type == "str":
            if not isinstance(value, str):
                raise ValueError("must be text")
            parsed: Any = value.strip()
        elif spec.data_type == "int":
            if isinstance(value, bool):
                raise ValueError("must be an integer")
            if isinstance(value, int):
                parsed = value
            elif isinstance(value, float) and value.is_integer():
                parsed = int(value)
            elif isinstance(value, str) and value.strip().lstrip("+-").isdigit():
                parsed = int(value.strip())
            else:
                raise ValueError("must be an integer")
        elif spec.data_type == "float":
            if isinstance(value, bool):
                raise ValueError("must be a number")
            parsed = float(value.strip() if isinstance(value, str) else value)
        elif spec.data_type == "bool":
            if isinstance(value, bool):
                parsed = value
            elif isinstance(value, str) and value.strip().lower() in {"true", "false"}:
                parsed = value.strip().lower() == "true"
            else:
                raise ValueError("must be true or false")
        elif spec.data_type == "list[str]":
            if not isinstance(value, str):
                raise ValueError("must be comma-separated text")
            parsed = [item.strip() for item in value.split(",")]
            if not parsed or any(not item for item in parsed):
                raise ValueError("must not contain empty list members")
            if spec.key.endswith("phase_names") and len(parsed) != len(set(parsed)):
                raise ValueError("must not contain duplicate phase names")
        else:  # pragma: no cover - FieldSpec constrains this at type-check time.
            raise ValueError(f"has unsupported data type {spec.data_type}")
    except (TypeError, ValueError) as exc:
        raise ConfigError(f"[{location}] {spec.key} {exc}.") from exc

    if spec.allowed_values and str(parsed) not in spec.allowed_values:
        allowed = ", ".join(spec.allowed_values)
        raise ConfigError(f"[{location}] {spec.key} must be one of: {allowed}.")
    return parsed


def _metadata_values(spec: FieldSpec) -> tuple[Any, ...]:
    return (
        spec.section,
        spec.name,
        spec.data_type,
        spec.unit,
        "Yes" if spec.required else "No",
        spec.required_when,
        ",".join(spec.allowed_values),
        spec.description,
        spec.chinese_notes,
    )


def _display_value(value: Any) -> Any:
    if isinstance(value, list):
        return ",".join(str(item) for item in value)
    return deepcopy(value)


def _style_data_sheet(sheet: Any, headers: tuple[str, ...], input_columns: Iterable[int]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
    for column in input_columns:
        letter = sheet.cell(1, column).column_letter
        sheet.column_dimensions[letter].width = 18
    for column in range(1, len(headers) + 1):
        letter = sheet.cell(1, column).column_letter
        if sheet.column_dimensions[letter].width == 13.0:
            sheet.column_dimensions[letter].width = 16


def _add_validation(sheet: Any, row: int, spec: FieldSpec) -> None:
    if spec.allowed_values:
        formula = '"' + ",".join(spec.allowed_values) + '"'
    elif spec.data_type == "bool":
        formula = '"TRUE,FALSE"'
    else:
        return
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Choose a value from the list."
    validation.errorTitle = "Invalid value"
    sheet.add_data_validation(validation)
    validation.add(sheet.cell(row, 5))


def _iter_dynamic_specs(data: dict[str, Any]) -> list[tuple[str, FieldSpec]]:
    found: list[tuple[str, FieldSpec]] = []

    def visit(value: Any, prefix: str = "") -> None:
        if isinstance(value, dict):
            for child_key, child_value in value.items():
                key = f"{prefix}.{child_key}" if prefix else child_key
                visit(child_value, key)
            return
        spec = field_spec_for_key(prefix)
        if spec is not None and spec.pattern is not None:
            found.append((prefix, spec))

    visit(data)
    return sorted(found, key=lambda item: item[0])


def _write_scalar_rows(sheet: Any, data: dict[str, Any], template: bool) -> None:
    rows: list[tuple[str, FieldSpec]] = [
        (spec.key, spec) for spec in FIELD_SPECS if spec.pattern is None
    ]
    if not template:
        rows.extend(_iter_dynamic_specs(data))

    for row_number, (key, spec) in enumerate(rows, start=2):
        value = spec.template_default if template else _get_dotted(data, key)
        default_value = None if value is _MISSING else _display_value(value)
        section, name, data_type, unit, required, applies_when, allowed, description, notes = _metadata_values(spec)
        sheet.append(
            [
                section,
                key,
                name,
                default_value,
                None,
                data_type,
                unit,
                required,
                applies_when,
                allowed,
                description,
                notes,
            ]
        )
        sheet.cell(row_number, 5).fill = _INPUT_FILL
        _add_validation(sheet, row_number, spec)
    sheet.auto_filter.ref = f"A1:L{sheet.max_row}"
    widths = (16, 48, 28, 18, 18, 12, 12, 12, 22, 26, 38, 28)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width


def _create_workbook(data: dict[str, Any], template: bool) -> Workbook:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions["A1"] = "Configuration workbook instructions"
    instructions["A2"] = "User Value overrides Default Value. Formula cells are not supported."
    instructions["A3"] = "Blank required values are invalid; nullable values become None; optional values are omitted."
    instructions.column_dimensions["A"].width = 100

    config_sheet = workbook.create_sheet("Config Parameters")
    _style_data_sheet(config_sheet, CONFIG_HEADERS, (4, 5))
    _write_scalar_rows(config_sheet, data, template)

    for sheet_name, headers in _SHEET_HEADERS.items():
        if sheet_name == "Config Parameters":
            continue
        sheet = workbook.create_sheet(sheet_name)
        _style_data_sheet(sheet, headers, range(1, len(headers)))
    return workbook


def _require_xlsx_path(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ConfigError(f"XLSX configuration files must use the .xlsx extension: {path}")


def save_xlsx(data: dict[str, Any], path: str | Path, *, template: bool = False) -> Path:
    destination = Path(path)
    _require_xlsx_path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = _create_workbook(data, template)
    workbook.save(destination)
    return destination


def create_xlsx_template(path: str | Path) -> Path:
    return save_xlsx({}, path, template=True)


def _validate_workbook_layout(workbook: Any) -> None:
    if set(workbook.sheetnames) != set(DATA_SHEETS) or len(workbook.sheetnames) != len(DATA_SHEETS):
        expected = ", ".join(DATA_SHEETS)
        raise ConfigError(f"Workbook must contain exactly these sheets: {expected}.")
    for sheet_name, headers in _SHEET_HEADERS.items():
        actual = tuple(cell.value for cell in workbook[sheet_name][1])
        if actual != headers:
            raise ConfigError(f"[{sheet_name}!A1] headers do not match the workbook contract.")


def _validate_metadata(sheet: Any, row: int, spec: FieldSpec) -> None:
    expected = _metadata_values(spec)
    columns = (1, 3, 6, 7, 8, 9, 10, 11, 12)
    for column, expected_value in zip(columns, expected):
        actual = sheet.cell(row, column).value
        normalized_actual = "" if _is_blank(actual) else actual
        normalized_expected = "" if _is_blank(expected_value) else expected_value
        if normalized_actual != normalized_expected:
            location = f"Config Parameters!{sheet.cell(row, column).coordinate}"
            raise ConfigError(f"[{location}] metadata does not match registry for {spec.key}.")


def _apply_blank_mode(data: dict[str, Any], spec: FieldSpec, location: str) -> None:
    if spec.blank_mode == BlankMode.ERROR:
        raise ConfigError(f"[{location}] {spec.key} is required.")
    if spec.blank_mode == BlankMode.NONE:
        _set_dotted(data, spec.key, None)


def _validate_dynamic_phase_names(data: dict[str, Any]) -> None:
    for key, value in _iter_dynamic_specs(data):
        del value
        tokens = key.split(".")
        channel, phase = tokens[1], tokens[3]
        names = _get_dotted(data, f"channels.{channel}.phase_names", [])
        if not isinstance(names, list) or phase not in names:
            raise ConfigError(
                f"[Config Parameters] {key} names phase {phase!r}, which is not configured for channels.{channel}.phase_names."
            )


def load_xlsx(path: str | Path, *, allow_partial: bool = False) -> dict[str, Any]:
    source = Path(path)
    _require_xlsx_path(source)
    try:
        workbook = load_workbook(source, data_only=False)
    except OSError as exc:
        raise ConfigError(f"Unable to load workbook: {source}") from exc
    _validate_workbook_layout(workbook)
    sheet = workbook["Config Parameters"]
    data: dict[str, Any] = {}
    seen: set[str] = set()

    for row in range(2, sheet.max_row + 1):
        key_cell = sheet.cell(row, 2)
        key = key_cell.value
        key_location = f"Config Parameters!{key_cell.coordinate}"
        if _is_blank(key):
            raise ConfigError(f"[{key_location}] Config Key is required.")
        if not isinstance(key, str):
            raise ConfigError(f"[{key_location}] Config Key must be text.")
        key = key.strip()
        if key in seen:
            raise ConfigError(f"[{key_location}] duplicate Config Key: {key}.")
        seen.add(key)
        spec = field_spec_for_key(key)
        if spec is None:
            raise ConfigError(f"[{key_location}] unknown Config Key: {key}.")
        _validate_metadata(sheet, row, spec)

        user_cell = sheet.cell(row, 5)
        default_cell = sheet.cell(row, 4)
        _reject_formula(default_cell.value, f"Config Parameters!{default_cell.coordinate}")
        _reject_formula(user_cell.value, f"Config Parameters!{user_cell.coordinate}")
        value_cell = user_cell if not _is_blank(user_cell.value) else default_cell
        location = f"Config Parameters!{value_cell.coordinate}"
        if _is_blank(value_cell.value):
            _apply_blank_mode(data, spec, location)
            continue
        _set_dotted(data, key, _parse_scalar(spec, value_cell.value, location))

    required_rows = {spec.key for spec in FIELD_SPECS if spec.pattern is None}
    if not allow_partial:
        missing = sorted(required_rows - seen)
        if missing:
            raise ConfigError(f"[Config Parameters] missing registry rows: {', '.join(missing)}.")
    _validate_dynamic_phase_names(data)
    return data
