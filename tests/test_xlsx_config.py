from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from obc_adc_designer.config import ConfigError, load_yaml
from obc_adc_designer.xlsx_config import DATA_SHEETS, load_xlsx, save_xlsx


class XlsxScalarConfigTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.root = Path(self.tempdir.name)
        self.cfg = load_yaml(ROOT / "config" / "pmp23607_default.yaml")

    def test_scalar_round_trip_and_layout(self) -> None:
        path = save_xlsx(self.cfg, self.root / "config.xlsx")
        loaded = load_xlsx(path)
        self.assertEqual(loaded["metadata"]["profile"], self.cfg["metadata"]["profile"])
        self.assertEqual(loaded["system"]["rated_power_w"], 2500.0)
        workbook = load_workbook(path, data_only=False)
        self.assertEqual(workbook.sheetnames, list(DATA_SHEETS))
        sheet = workbook["Config Parameters"]
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[-1], "Chinese Notes")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertIsNotNone(sheet.auto_filter.ref)

    def test_user_value_overrides_default_value(self) -> None:
        path = save_xlsx(self.cfg, self.root / "override.xlsx")
        workbook = load_workbook(path)
        sheet = workbook["Config Parameters"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 2).value == "system.rated_power_w":
                sheet.cell(row, 5).value = 3300
                break
        workbook.save(path)
        self.assertEqual(load_xlsx(path)["system"]["rated_power_w"], 3300.0)

    def test_scalar_rows_may_be_reordered(self) -> None:
        path = save_xlsx(self.cfg, self.root / "reordered.xlsx")
        workbook = load_workbook(path)
        sheet = workbook["Config Parameters"]
        first = [cell.value for cell in sheet[2]]
        last = [cell.value for cell in sheet[sheet.max_row]]
        for column, value in enumerate(last, start=1):
            sheet.cell(2, column).value = value
        for column, value in enumerate(first, start=1):
            sheet.cell(sheet.max_row, column).value = value
        workbook.save(path)
        self.assertEqual(
            load_xlsx(path)["metadata"]["profile"],
            self.cfg["metadata"]["profile"],
        )

    def test_formula_in_default_value_is_rejected_even_when_overridden(self) -> None:
        path = save_xlsx(self.cfg, self.root / "formula.xlsx")
        workbook = load_workbook(path)
        sheet = workbook["Config Parameters"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 2).value == "system.rated_power_w":
                sheet.cell(row, 4).value = "=1+1"
                sheet.cell(row, 5).value = 3300
                break
        workbook.save(path)
        with self.assertRaisesRegex(ConfigError, r"Config Parameters!D.*formula"):
            load_xlsx(path)
